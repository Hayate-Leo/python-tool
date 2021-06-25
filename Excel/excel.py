import openpyxl as opx


class Excel:
    def __init__(self, path):
        self.path = path
        self.read()

    def read(self):
        self.wb = opx.load_workbook(self.path)
        self.ws = self.wb.worksheets[0]

    def get_value_list(self, ws_2d):
        return([[cell.value for cell in row] for row in ws_2d])
    
    def get_list_2d(self, start_row, start_col, end_row, end_col):
        return self.get_value_list(self.ws.iter_rows(min_row=start_row,
                                                    max_row=end_row,
                                                    min_col=start_col,
                                                    max_col=end_col))
    
    def get_list_1d(self, start_row, col, end_row=False):
        return [row[col].value for row in self.ws.iter_rows(min_row=start_row, max_row=end_row)]

    def write_list_2d(self, l_2d, start_row, start_col):
        for y, row in enumerate(l_2d):
            for x, cell in enumerate(row):
                self.ws.cell(row=start_row + y,
                                column=start_col + x,
                                value=l_2d[y][x])

        self.wb.save(self.path)
    
    def write_list_1d(self, l_1d, start_row, start_col):
        for i, element in enumerate(l_1d):
            self.ws.cell(row=start_row + i, column=start_col, value=element)

        self.wb.save(self.path)
    

if __name__ == '__main__':
    excel = Excel('Excel/sample.xlsx')
    list_2d = excel.get_list_2d(2, 3, excel.ws.max_row-2, excel.ws.max_column-3)
    list_1d = excel.get_list_1d(2, 3)
    excel.write_list_2d(list_2d, 2, 10)
    excel.write_list_1d(list_1d, 20, 3)